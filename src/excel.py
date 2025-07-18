import openpyxl
from components import Components
from settings import Config
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, PatternFill
import conf

def set_border(ws, row, cols):

#	thin_border = Border(left=Side(style='thin'), 
#                   right=Side(style='thin'), 
#                   top=Side(style='thin'), 
#                   bottom=Side(style='thin'))

	cell_border = Border(top=Side(style='thin'))
	for col in cols:
		ws[col+row].border = cell_border

class XlsDocument:

	thin_left = Border(left=Side(style='thin'))
	thin_right = Border(right=Side(style='thin'))
	thin_left_top = Border(left=Side(style='thin'),
					top=Side(style='thin'))
	thin_right_top = Border(right=Side(style='thin'),
					top=Side(style='thin'))

	def __init__(self, components: Components):
		self.data = components.model.data
		self.components = components


	def save_in_xls(self):

		if self.data["lang"] == "eng":
			xlsx_template = Config.xlsx_template_eng
			sheet_template = Config.sheet_templates_eng
		else:
			xlsx_template = Config.xlsx_template_ita
			sheet_template = Config.sheet_templates_ita


		xlsx_template = "/usr/local/src/eurotherm/" + xlsx_template
		wb = openpyxl.load_workbook(xlsx_template)


		for ws_data in sheet_template:
			ws = wb[ws_data[0]]
			self.xls_headers(ws)
			self.room_breakdown(ws, ws_data)

		out = conf.spool+self.data["outfile"][:-4]+".xlsx"		
		wb.save(out)


	def xls_headers(self, ws):
	
		model = self.components.model
		no_collectors = len(model.collectors)

		# copy total area
		ws['B14'] =	ws['B18'] = model.area

		# copy total coverage
		ws['A14'] =	ws['A18'] = model.active_area


		record = self.components.panel_record
		total_full_panels = (record["full_classic"] +
							 record["full_hydro"] +
							 record["lux_classic"] +
						     record["lux_hydro"])
							
		total_split_panels = (record["split_classic"] +
							  record["split_hydro"])

		# copy total panels
		ws['D14'] = ws['D18'] = total_full_panels
		ws['E14'] = ws['E18'] = total_split_panels

		# copy number of feeds
		ws['F14'] = ws['F18'] = self.components.num_lines

		# copy number of collectors 
		ws['G14'] = ws['G18'] = str(no_collectors) 
		


	def room_breakdown(self, ws, ws_data):

		model = self.components.model

		warm_coef = ws_data[2]
		cool_coef = ws_data[3]

		ws['A22'] = str(ws_data[1])
		ws.row_dimensions[3].height = 32

		# header
		for i in range(65,85):
			if (67<=i<=78):
				ws.column_dimensions[chr(i)].width = 10
			ws[chr(i)+'23'].alignment = \
				Alignment(wrapText=True, 
					vertical ='center',
					horizontal ='center')

		self.rooms_header(ws)

		set_border(ws, '23', "ABCDEFGHIJKLMNOPQ")
		ws['A23'].border = XlsDocument.thin_left_top
		ws['B23'].border = XlsDocument.thin_left_top
		ws['G23'].border = XlsDocument.thin_left_top
		ws['L23'].border = XlsDocument.thin_left_top
		ws['O23'].border = XlsDocument.thin_left_top
		ws['Q23'].border = XlsDocument.thin_right_top

		model.processed.sort(key=lambda x: 
			(x.collector.zone_num, x.collector.number, x.pindex))

		for locale in model.locales:
			if (locale.collector == "" and locale.room.collector):
				locale.collector = locale.room.collector.name

		model.locales.sort(key=lambda x:
			(x.zone, x.collector, x.pindex))

		zone = 0
		index = 24
		number = -1
		for locale in model.locales:

			ws['A'+str(index)].border = XlsDocument.thin_left
			ws['B'+str(index)].border = XlsDocument.thin_left
			ws['G'+str(index)].border = XlsDocument.thin_left
			ws['L'+str(index)].border = XlsDocument.thin_left
			ws['O'+str(index)].border = XlsDocument.thin_left
			ws['Q'+str(index)].border = XlsDocument.thin_right

			# body
			for i in range(65,82):
				pos = chr(i)+str(index)
				ws[pos].alignment = \
					Alignment(wrapText=True, 
						vertical ='center',
						horizontal ='center')
				color = "D0D0D0"
				if (index%2==0 and i>65):
					color = "F0F0F0"
				ws[pos].fill = PatternFill(start_color=color, 
						fill_type = "solid")

			while (locale.zone>zone):
				zone += 1
				pos = 'A' + str(index)
				ws[pos] = "Zona %d" % zone
				set_border(ws,str(index), 'A')
			
			if (locale.collector != number):
				number = locale.collector
				set_border(ws, str(index), "BCDEFGHIJKLMNOPQ")
				ws['B'+str(index)].border = XlsDocument.thin_left_top
				ws['G'+str(index)].border = XlsDocument.thin_left_top
				ws['L'+str(index)].border = XlsDocument.thin_left_top
				ws['O'+str(index)].border = XlsDocument.thin_left_top
				ws['Q'+str(index)].border = XlsDocument.thin_right_top

			pos = 'B' + str(index)
			ws[pos] = locale.collector
			ws[pos].alignment = Alignment(horizontal='center')
			
			pos = 'C' + str(index)
			ws[pos] = locale.name
			if not locale.name.isnumeric():
				ws[pos].number_format = "@"
			else:
				ws[pos] = int(locale.name)
			ws[pos].alignment = Alignment(horizontal='center')

			pos = 'E' + str(index)
			ws[pos] = locale.room.area_m2()
			ws[pos].number_format = "0.0"

			if (locale.room.active_m2==0):
				index += 1
				continue

			pos = 'D' + str(index)
			ws[pos] = locale.active_m2
			ws[pos].number_format = "0.0"

			pos = 'F' + str(index)
			ws[pos] = locale.room.ratio
			ws[pos].number_format = "0.0%"

			pos = 'G' + str(index)
			ws[pos] = locale.lines
				
			if (locale.panel_record["full_classic"] +
				locale.panel_record["lux_classic"] +
				locale.panel_record["full_hydro"]   +
				locale.panel_record["lux_hydro"]	>0):
				pos = 'H' + str(index)
				ws[pos] = (locale.panel_record["full_classic"] +
						   locale.panel_record["lux_classic"] +
						   locale.panel_record["full_hydro"]  +
						   locale.panel_record["lux_hydro"])

			if (locale.panel_record["split_classic"]+
				locale.panel_record["split_hydro"]>0):
				pos = 'I' + str(index)
				ws[pos] = (locale.panel_record["split_classic"] + 
						   locale.panel_record["split_hydro"])

			if (locale.panel_record["half_classic"]+
				locale.panel_record["half_hydro"]>0):
				pos = 'J' + str(index)
				ws[pos] = (locale.panel_record["half_classic"] + 
						   locale.panel_record["half_hydro"])

			if (locale.panel_record["quarter_classic"]+
				locale.panel_record["quarter_hydro"]>0):
				pos = 'K' + str(index)
				ws[pos] = (locale.panel_record["quarter_classic"] + 
						   locale.panel_record["quarter_hydro"])


			# heating 
			pos = 'L' + str(index)
			ws[pos] = radiated = locale.active_m2 * warm_coef
			ws[pos].number_format = "0"

			pos = 'M' + str(index)
			ws[pos] = output = radiated * 1.1
			ws[pos].number_format = "0"

			pos = 'N' + str(index)
			ws[pos] = 3.6*output/(4.186*ws['K14'].value)
			ws[pos].number_format = "0"

			# cooling
			pos = 'O' + str(index)
			ws[pos] = absorbed = locale.active_m2 * cool_coef
			ws[pos].number_format = "0"

			pos = 'P' + str(index)
			ws[pos] = output = absorbed * 1.1
			ws[pos].number_format = "0"

			pos = 'Q' + str(index)
			ws[pos] = 3.6*output/(4.186*ws['K18'].value)
			ws[pos].number_format = "0"


			#ws[pos_area].number_format = "0.00"
			index += 1

		set_border(ws, str(index), "ABCDEFGHIJKLMNOPQ")


	def rooms_header(self, ws):

		ws.row_dimensions[23].height = 32

		if self.data["lang"] == "ita":
			ws['A23'] = "Zona"
			ws['B23'] = "Collettore"
			ws['C23'] = "Stanza"

			ws['D23'] = "Attiva\n[m2]"
			ws['E23'] = "Area\n[m2]"
			ws['F23'] = "% cop."
			ws['G23'] = "linee"

			ws['H23'] = "Pannelli\n200x120"
			ws['I23'] = "Pannelli\n200x60"
			ws['J23'] = "Pannelli\n100x120"
			ws['K23'] = "Pannelli\n100x60"
				
			#ws.column_dimensions['M'].width = 2

			ws['L22'] = 'Riscaldamento'
			ws['L23'] = "Q, resa\n[W]"
			ws['M23'] = "Q, tot\n[W]"
			ws['N23'] = "Portata\n[kg/h]"

			#ws.column_dimensions['Q'].width = 2

			ws['O22'] = 'Raffrescamento'
			ws['O23'] = "Q, resa\n[W]"
			ws['P23'] = "Q, tot\n[W]"
			ws['Q23'] = "Portata\n[kg/h]"

			ws['A23'] = "Zona"
			ws['B23'] = "Collettore"
			ws['C23'] = "Stanza"

			ws['D23'] = "Attiva\n[m2]"
			ws['E23'] = "Area\n[m2]"
			ws['F23'] = "% cop."
			ws['G23'] = "linee"

			ws['H23'] = "Pannelli\n200x120"
			ws['I23'] = "Pannelli\n200x60"
			ws['J23'] = "Pannelli\n100x120"
			ws['K23'] = "Pannelli\n100x60"
				
			#ws.column_dimensions['M'].width = 2

			ws['L22'] = 'Riscaldamento'
			ws['L23'] = "Q, resa\n[W]"
			ws['M23'] = "Q, tot\n[W]"
			ws['N23'] = "Portata\n[kg/h]"

			#ws.column_dimensions['Q'].width = 2

			ws['O22'] = 'Raffrescamento'
			ws['O23'] = "Q, resa\n[W]"
			ws['P23'] = "Q, tot\n[W]"
			ws['Q23'] = "Portata\n[kg/h]"

		if self.data["lang"] == "eng":
			ws['A23'] = "Zone"
			ws['B23'] = "Collector"
			ws['C23'] = "Room"

			ws['D23'] = "Active\n[m2]"
			ws['E23'] = "Area\n[m2]"
			ws['F23'] = "% cov."
			ws['G23'] = "lines"

			ws['H23'] = "Panels\n200x120"
			ws['I23'] = "Panels\n200x60"
			ws['J23'] = "Panels\n100x120"
			ws['K23'] = "Panels\n100x60"
				
			#ws.column_dimensions['M'].width = 2

			ws['L22'] = 'Heating'
			ws['L23'] = "Q, yield\n[W]"
			ws['M23'] = "Q, tot\n[W]"
			ws['N23'] = "Flow\n[kg/h]"

			#ws.column_dimensions['Q'].width = 2

			ws['O22'] = 'Cooling'
			ws['O23'] = "Q, yield\n[W]"
			ws['P23'] = "Q, tot\n[W]"
			ws['Q23'] = "Flow\n[kg/h]"

			ws['A23'] = "Zone"
			ws['B23'] = "Collector"
			ws['C23'] = "Room"

			ws['D23'] = "Active\n[m2]"
			ws['E23'] = "Area\n[m2]"
			ws['F23'] = "% cov."
			ws['G23'] = "lines"

			ws['H23'] = "Panels\n200x120"
			ws['I23'] = "Panels\n200x60"
			ws['J23'] = "Panels\n100x120"
			ws['K23'] = "Panels\n100x60"
				
			#ws.column_dimensions['M'].width = 2

			ws['L22'] = 'Heating'
			ws['L23'] = "Q, yield\n[W]"
			ws['M23'] = "Q, tot\n[W]"
			ws['N23'] = "Flow\n[kg/h]"

			#ws.column_dimensions['Q'].width = 2

			ws['O22'] = 'Cooling'
			ws['O23'] = "Q, yield\n[W]"
			ws['P23'] = "Q, tot\n[W]"
			ws['Q23'] = "Flow\n[kg/h]"
