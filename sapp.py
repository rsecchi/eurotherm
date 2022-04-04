#!/usr/local/bin/python3

import ezdxf
import openpyxl
from openpyxl.styles.borders import Border, Side
import os.path
from math import ceil, floor, sqrt
from tkinter import *
from tkinter import filedialog
from tkinter.messagebox import askyesno

# Parameter settings (values in cm)
default_scale = 1          # multiplier to transform in cm (scale=100 if the drawing is in m)
default_tolerance    = 1   # ignore too little variations
default_width_doga   = 2
default_space_omega  = 60
default_croc_first   = 15
default_croc_second  = 60
default_croc_maxd    = 120
default_croc_tol     = 10
default_cut_size     = 800

extra_len    = 20

default_zone_cost    = 1   # equivalent cost of a zone in m2 material

default_x_font_size  = 20
default_y_font_size  = 30

default_input_layer = 'AREE_SAPP'
layer_text   = 'Eurotherm_text'
layer_box    = 'Eurotherm_box'
layer_croc   = 'Eurotherm_crocodile'
layer_omega   = 'Eurotherm_omega'

text_color = 7
box_color = 8
croc_color = 9
omega_color = 10

xlsx_template = 'template.xlsx'
sheet_template = 'BoM'

alphabet = {
	' ': [],
	'A': [12,3,1,5,14,8,6],
	'B': [0,1,5,7,6,7,11,13,12,0],
	'C': [2,1,3,9,13,14],
	'D': [0,1,5,11,13,12,0],
	'E': [2,0,6,7,6,12,14],
	'F': [2,0,6,7,6,12],
	'G': [2,1,3,9,13,14,8,7],
	'H': [0,12,6,8,2,14],
	'I': [1,13],
	'J': [9,13,1,0,1,2],
	'K': [0,6,2,6,12,6,14],
	'L': [0,12,14],
	'M': [12,0,7,2,14],
	'N': [12,0,14,2],
	'O': [3,1,5,11,13,9,3],
	'P': [12,0,1,5,7,6],
	'Q': [13,11,5,1,3,9,13,14,13,10],
	'R': [12,0,1,5,7,6,7,11,14],
	'S': [12,13,11,3,1,2],
	'T': [0,2,1,13],
	'U': [0,9,13,11,2],
	'V': [0,13,2],
	'W': [0,13,2],
	'X': [0,14,7,2,12],
	'Y': [0,7,2,7,13],
	'Z': [0,2,12,14],
	'0': [1,3,9,13,11,5,1],
	'1': [3,1,13,12,14],
	'2': [3,1,5,12,14],
	'3': [0,1,5,7,6,7,11,13,12],
	'4': [0,6,8,2,14],
	'5': [2,0,6,7,11,13,12],
	'6': [2,1,3,9,13,11,7,6],
	'7': [0,2,12],
	'8': [3,1,5,9,12,14,11,3],
	'9': [8,6,3,1,5,11,13,12]
}

def pletter(letter,pos,scale):
	path = alphabet[letter]
	poly = []
	for p in path:
		x = p % 3
		y = 4 - (p//3)
		poly.append((pos[0]+x*scale[0], pos[1]+y*scale[1]))
	return poly

def writedxf(msp, strn, pos, scale):

	slen = len(strn)*scale[0]*3
	pos = (pos[0]-slen/2, pos[1]-scale[1]*2)
	
	for l in strn:
		poly = pletter(l, pos, scale)
		pl = msp.add_lwpolyline(poly)
		pl.dxf.layer = layer_text
		pos = (pos[0]+scale[0]*3, pos[1])

def spread(a,b,size):
	l = []
	n = ceil(abs(a-b)/size) - 1
	L = abs(a-b)/(n+1)
	for i in range(0,n):
		l.append((i+1)*L+a)
	return l

def center(a,b,size):
	l = []
	n = ceil(abs(b-a)/size) - 1
	L = (abs(b-a) - size*n)/2
	for i in range(0,n+1):
		l.append(a+L+i*size)	
	return l


def intersects(p, x):
		
	ints = []
	eps = 1e-6
		
	for i in range(0, len(p)-1):
		vx, vy = p[i+1][0]-p[i][0], p[i+1][1]-p[i][1]

		if (p[i+1][0]==x or p[i][0]==x):
				x += eps

		if (p[i][0] < p[i+1][0]):
			x0, y0 = p[i][0], p[i][1]
			x1, y1 = p[i+1][0], p[i+1][1]
		else:
			x0, y0 = p[i+1][0], p[i+1][1]
			x1, y1 = p[i][0], p[i][1]
					
		if (x0<x and x<x1):
			delta = x1 - x0
			py = (y0*(x1-x) - y1*(x0-x))/delta
			ints.append(py)
	
	return sorted(ints)


class Split:
	
	def __init__(self, ax, bx, lu, ld, ru, rd):
		self.ax = ax
		self.bx = bx
		self.lu = lu
		self.ld = ld
		self.ru = ru
		self.rd = rd

	def base(self):
		return abs(self.bx - self.ax)

	def area(self):
		l1 = abs(self.lu[0] - self.ld[0])
		l2 = abs(self.ru[0] - self.rd[0])
		b = self.base()
		return (l1+l2)*b/20000


class Cluster:
	def __init__(self, box, x):
		self.x = x
		self.ax = box.ax
		self.bx = box.bx
		self.ay = box.ay
		self.by = box.by
		if (box.bx == x):
			self.boxesl = [box]
			self.boxesr = []
		else:
			self.boxesl = []
			self.boxesr = [box]

	def print(self):

		print("x=%.2f [%.2f %.2f %.2f %.2f]:" % (self.x, self.ax, self.bx, self.ay, self.by))

		print("LB:", end='')
		for box in self.boxesl:
			print("box(%d) [%.2f %.2f %.2f %.2f]   " % (box.id, box.ax, box.bx, box.ay, box.by), end='')
		print("")

		print("RB:", end='')
		for box in self.boxesr:
			print("box(%d) [%.2f %.2f %.2f %.2f]   " % (box.id, box.ax, box.bx, box.ay, box.by), end='')
		print("\n")

		
	def bb_area(self):
		ax = self.ax
		ay = self.ay
		bx = self.bx
		by = self.by
		return abs(bx-ax)*abs(by-ay)/10000

	def area(self):
		area_ls = 0
		for box in self.boxesl:
			area_ls += box.area()
		area_rs = 0
		for box in self.boxesr:
			area_rs += box.area()

		return area_ls + area_rs
		
	def scrap(self):
		s =  self.bb_area() - self.area()
		return s
		

	def append(self, box):

		if (box.ax != self.x and box.bx != self.x):
			return

		if (box.ax < self.ax):
			self.ax = box.ax

		if (box.bx > self.bx):
			self.bx = box.bx

		if (box.ay < self.ay):
			self.ay = box.ay

		if (box.by > self.by):
			self.by = box.by

		if (box.ax == self.x):
			self.boxesr.append(box)
		
		if (box.bx == self.x):
			self.boxesl.append(box)


	def check_append(self, box):

		if (box.ax==self.x or box.bx==self.x):
			if (not (box.ay>self.by or box.by<self.ay)):
				self.append(box)
				return True
		return False

	def can_merge(self, cluster):
		if (self.x == cluster.x):
			if (self.ay<self.by or self.by>self.ay):
				return True
		return False

	@classmethod
	def _get_one(cls, clusters):
		flag = False
		for cl1 in clusters:
			for cl2 in clusters:
				if (cl2==cl1):
					continue
				if (cl2.can_merge(cl1)):
					return (cl1, cl2)

	@classmethod
	def merge(cls, clusters):
		
		while( (cc := cls._get_one(clusters)) != None):

			cc[0].boxesr += cc[1].boxesr
			cc[0].boxesl += cc[1].boxesl

			if (cc[0].ay > cc[1].ay):
				cc[0].ay = cc[1].ay
			if (cc[0].by < cc[1].by):
				cc[0].by = cc[1].by
			if (cc[0].ax > cc[1].ax):
				cc[0].ax = cc[1].ax
			if (cc[0].bx < cc[1].bx):
				cc[0].bx = cc[1].bx

			clusters.remove(cc[1])


class Croc:
	def __init__(self, x1, x2, y):
		self.x1 = x1
		self.x2 = x2
		self.y  = y 

	def fit(self, y1, y2):
		tol = croc_tol
		return (self.y>=y1 and self.y<=y2)


	def extend_right(self, x2):
		self.x2 = x2

	def len(self):
		return abs(self.x2-self.x1)/100

class Zone:
	
	ident = 0

	def __init__(self, ax, ay, bx, by, split):
		self.id = Zone.ident
		Zone.ident += 1
		self.ax = min(ax, bx)
		self.bx = max(ax, bx)
		self.ay = min(ay, by)
		self.by = max(ay, by)
		self.TR = (max(ax, bx), max(ay, by))
		self.TL = (min(ax, bx), max(ay, by))
		self.BR = (max(ax, bx), min(ay, by))
		self.BL = (min(ax, bx), min(ay, by))
		if (type(split) == list):
			self.splits = split
		else:
			self.splits = [split]

	def pline(self, orient):
		ax = self.ax; bx = self.bx
		ay = self.ay; by = self.by
		if (orient == 0):
			return [(ax,ay),(ax,by),(bx,by),(bx,ay),(ax,ay)]
		else:	
			return [(ay,ax),(by,ax),(by,bx),(ay,bx),(ay,ax)]

	def extend_right(self, coord):
		self.TR = (coord, self.TR[1])
		self.BR = (coord, self.BR[1])
		self.bx = coord

	def extend_crocs(self, crocs, y1, y2):
		crocs_list = list()
		for c in crocs:
			print("BOX",self.ax,y1,y2)
			if (c.x2 == self.ax and c.y>=y1 and c.y<=y2):
				print("exte",c.x1,c.x2,c.y)
				c.extend_right(self.bx)
				crocs_list.append(c)
		return crocs_list

	def draw_box(self, msp, orient):
		pl = msp.add_lwpolyline(self.pline(orient))
		pl.dxf.layer = layer_box

	def draw_text(self, msp, ind, subind, orient):
		scale = (x_font_size/2, y_font_size/4)
		if (subind == -1):
			text = 'ZONE ' + str(ind)
		else:
			text = 'ZONE ' + str(ind) + chr(65+(subind-1)%26) 
		if (orient==0):
			pos = ((self.ax+self.bx)/2, (self.ay+self.by)/2)
		else:
			pos = ((self.ay+self.by)/2, (self.ax+self.bx)/2)
	
		writedxf(msp, text, pos, scale)

	def area(self):
		return (self.bx-self.ax)*(self.by-self.ay)/10000

	def splitarea(self):
		totsplit = 0
		for split in self.splits:
			totsplit = totsplit + split.area()
		return totsplit

	def scrap(self):
		return self.area() - self.splitarea()


	def overlap(self, box):
		c1 = self.ax >= box.bx
		c2 = self.bx <=  box.ax
		c3 = self.ay >= box.by
		c4 = self.by <= box.ay

		return not(c1 or c2 or c3 or c4)

	def overlap_top(self, box):
		c1 = self.ax >= box.ax
		c2 = self.bx <= box.bx
		c3 = self.ay > box.ay and self.ay < box.by
		c4 = self.by > box.by
		return c1 and c2 and c3 and c4

	def overlap_bottom(self, box):
		c1 = self.ax >= box.ax
		c2 = self.bx <= box.bx
		c3 = self.ay < box.ay
		c4 = self.by > box.ay and self.by < box.by
		return c1 and c2 and c3 and c4

	def overlap_left(self, box):
		c1 = self.ay >= box.ay
		c2 = self.by <= box.by
		c3 = self.ax < box.ax 
		c4 = self.bx > box.ax and self.bx < box.bx
		return c1 and c2 and c3 and c4

	def overlap_right(self, box):
		c1 = self.ay >= box.ay
		c2 = self.by <= box.by
		c3 = self.ax > box.ax and self.ax < box.bx
		c4 = self.bx > box.bx
		return c1 and c2 and c3 and c4
		 

class Room:

	index = 1

	def __init__(self, poly):

		self.index = Room.index
		Room.index = Room.index + 1
		self.ignore = False
		self.errorstr = ""

		# Scale points to get cm
		color = poly.dxf.color
		if (color>3 or color<2):
			self.ignore = True
			self.errorstr = "WARNING: color=%d not supported, " % color
			self.errorstr += "ignoring Zone %d\n" % self.index
			return

		tol = tolerance
		self.orient = 0
		self.boxes = list()
		self.coord = list()
		self.crocs = list()
		self.omegas = list()

		self.points = list(poly.vertices())	

		# Add a final point to closed polylines
		p = self.points
		if (poly.is_closed):
			self.points.append((p[0][0], p[0][1]))

		# Check if the polyine is open with large final gap
		n = len(p)-1
		if (abs(p[0][0]-p[n][0])>tol or abs(p[0][1]-p[n][1])>tol):
			self.errorstr = "Error: open polyline in Zone %d \n" % self.index 
			self.ignore = True
			return

		# Straighten up walls
		finished = False
		while not finished:
			for i in range(1, len(p)):
				if (abs(p[i][0]-p[i-1][0]) < tol):
					p[i] = (p[i-1][0], p[i][1])

				if (abs(p[i][1]-p[i-1][1]) < tol):
					p[i] = (p[i][0], p[i-1][1])

			if (p[0] == p[n]):
				finished = True
			else:
				p[0] = p[n]

		# Projections of coordinates on x and y
		self.xcoord = sorted(set([p[0] for p in self.points]))	
		self.ycoord = sorted(set([p[1] for p in self.points]))	

		# Mirror if polyline is green
		if (poly.dxf.color==3):
			self.orient = 1
			for i in range(0,len(self.points)):
				self.points[i] = (self.points[i][1], self.points[i][0])
			self.xcoord, self.ycoord = self.ycoord, self.xcoord

		self.get_boxes()
		#self.get_crocs()
		self.simple_crocs()
		self.get_omegas()


	def area(self):
		a = 0
		p = self.points
		for i in range(0, len(p)-1):
			a += (p[i+1][0]-p[i][0])*(p[i+1][1] + p[i][1])/2
		return abs(a/10000)

	def perimeter(self):
		p = self.points
		d = 0
		for i in range(0, len(p)-1):
			d += sqrt( pow(p[i+1][0]-p[i][0], 2) + pow(p[i+1][1]-p[i][1], 2))
		return d

	# Building Room

	# This function divides a polyline into boxes
	# and returns the list of boxes
	def get_boxes(self):	

		tol = tolerance
		boxes = self.boxes
		p = self.points
		xc = self.xcoord

		for i in range(0,len(xc)-1):
			ax = xc[i]
			bx = xc[i+1]
			mid = (ax + bx)/2

			lr = list()
			ll = list()

			for j in range(0, len(p)-1):
				if (p[j][0] < p[j+1][0]):
					x0, y0 = p[j][0], p[j][1]
					x1, y1 = p[j+1][0], p[j+1][1]
				else:
					x0, y0 = p[j+1][0], p[j+1][1]
					x1, y1 = p[j][0], p[j][1]
					
				if (x0<mid and mid<x1):
					delta = x1 - x0
					ay = (y0*(x1-ax) - y1*(x0-ax))/delta
					by = (y0*(x1-bx) - y1*(x0-bx))/delta
					ll.append((ay,j))
					lr.append((by,j))

			ll.sort()
			lr.sort()

			for j in range(0,len(ll),2):
				ay = min(ll[j][0], lr[j][0])
				by = max(ll[j+1][0], lr[j+1][0])

				split = Split(ax, bx, ll[j], ll[j+1], lr[j], lr[j+1])
				boxes.append(Zone(ax,ay,bx,by, split))

		# collate splits
		while (self.collate_splits()): pass

		# minimize scrap
		to_add = []
		to_del = []
		for box in boxes:
			area = box.area()
			scrap = box.scrap()
			m = sqrt(scrap/zone_cost)
			n = max(floor(m), 1)		
			if ( n*zone_cost < scrap/(n+1) ):
				n = n + 1

			if (n>1):
				# instantiate new blocks
				ax = box.ax
				bx = box.bx
				d = (bx - ax)/n
				lu = box.splits[0].lu[0]
				ld = box.splits[0].ld[0]
				du = (box.splits[0].ru[0] - lu)/n
				dd = (box.splits[0].rd[0] - ld)/n
				
				j = box.splits[0].lu[1]
				k = box.splits[0].ld[1]

				for i in range(0,n):
					ru = lu + du
					rd = ld + dd
					split = Split(ax, ax+d, (lu,j), (ld,k), (ru,j), (rd,k))
					ay = min(lu, ld, ru, rd)
					by = max(lu, ld, ru, rd)
					ax_next = ax + d
					if (i<n-1):
						zs = Zone(ax, ay, ax_next, by, split)
					else:
						zs = Zone(ax, ay, bx, by, split)
					to_add.append(zs)
					ax = ax_next
					lu = lu + du
					ld = ld + dd 


				to_del.append(box)

		self.boxes = boxes + to_add
		for box in to_del:
			self.boxes.remove(box)



		# optimize
		done = False
		while(not done):

			self.get_clusters()
			mincost = 0
			for c in self.clusters:
				ay = c.ay
				by = c.by

				c.boxesr.sort(key=lambda a: a.bx)
				c.boxesl.sort(key=lambda a: a.ax, reverse=True)

				L = len(c.boxesl)
				R = len(c.boxesr)

				for i in range(0,R+1):
					for j in range(0,L+1):

						if (i+j<2):
							continue

						w = (R-i) + (L-j) + 1
						boxes = c.boxesr[0:i] + c.boxesl[0:j]

						# bounding boxes
						ax = boxes[0].ax
						bx = boxes[0].bx

						ta = 0
						for b in boxes:
							ta += b.area()
							if (b.ax<ax): ax=b.ax
							if (b.bx>bx): bx=b.bx

						scrap = (bx-ax)*(by-ay)/10000 - ta
						cost = scrap - zone_cost * w

						if (cost < mincost):
							mincost = cost
							axb = ax
							bxb = bx
							ayb = ay
							byb = by
							bestr = c.boxesr[0:i]
							bestl = c.boxesl[0:j]
							otherbr = c.boxesr[i:R]
							otherbl = c.boxesl[j:L]
							bestc = c

			#self.print_clusters()
			#bestc.print()

			if (mincost<0):
				# update zone partitioning
				
				splits = []
				for box in bestl:
					splits += box.splits
				for box in bestr:
					splits += box.splits

				for box in otherbl:
					zone = Zone(box.ax, box.ay, axb, box.by, box.splits)
					self.boxes.append(zone)

				for box in otherbr:
					zone = Zone(bxb, box.ay, box.bx, box.by, box.splits)
					self.boxes.append(zone)

				zone = Zone(axb, ayb, bxb, byb, splits)
				self.boxes.append(zone)

				for box in bestc.boxesr:
					self.boxes.remove(box)

				for box in bestc.boxesl:
					self.boxes.remove(box)

				# eliminate lateral overlaps
				for box in self.boxes:
					if (box != zone and box.overlap(zone)):
						if (box.overlap_top(zone)):
							box.ay = zone.by

						if (box.overlap_bottom(zone)):
							box.by = zone.ay

						if (box.overlap_left(zone)):
							box.bx = zone.ax

						if (box.overlap_right(zone)):
							box.ax = zone.bx

			else:
				done = True

		if (tcuts == 1):
			self.transversal_cuts()

		# renumber zones
		number = 1
		for box in self.boxes:
			box.number = number
			number += 1
	
	def transversal_cuts(self):

		to_remove = []
		for box in self.boxes:
			if (box.by-box.ay > cut_size):
				y0 = box.ay
				for y1 in spread(box.ay, box.by, cut_size):
					self.boxes.append(Zone(box.ax,y0,box.bx,y1, box.splits))
					y0 = y1

				self.boxes.append(Zone(box.ax,y0,box.bx,box.by, box.splits))
				to_remove.append(box)

		for box in to_remove:
			self.boxes.remove(box)

	def print_boxes(self):
		print("LIST OF %d BOXES: " % len(self.boxes), end='')
		for box in self.boxes:
			print("[%d]" % box.id, end='')
		print("")
			

	def print_clusters(self):		
		print("START CLUSTERS -------------------------")
		for cluster in self.clusters:
			cluster.print()
		print('END CLUSTERS   =========================')

	def get_clusters(self):
		clusters = self.clusters = []
		for box in self.boxes:

			lhs = False
			rhs = False

			for cluster in clusters:
				if (box.ax == cluster.x and rhs == False):
					if (cluster.check_append(box)):
						rhs = True

				if (box.bx == cluster.x and lhs == False):
					if (cluster.check_append(box)):
						lhs = True

			if (lhs == False):
				clusters.append(Cluster(box, box.bx))

			if (rhs == False):
				clusters.append(Cluster(box, box.ax))


		Cluster.merge(clusters)
				

	def collate_splits(self):
		for b in self.boxes:
			for q in self.boxes:
				if (b.ax == q.bx and b.splits[0].lu[1] == q.splits[0].ru[1] and
					                 b.splits[0].ld[1] == q.splits[0].rd[1]):
					lu = q.splits[0].lu 
					ld = q.splits[0].ld
					ru = b.splits[0].ru
					rd = b.splits[0].rd
					split = Split(q.ax, b.bx, lu, ld, ru, rd)
					top = max(q.by, b.by)
					btm = min(q.ay, b.ay)
					z = Zone(q.ax, btm, b.bx, top, split)
					self.boxes.remove(b)
					self.boxes.remove(q)
					self.boxes.append(z)
					return True

		return False


	def get_crocs(self):
		
		q = croc_first
		t = croc_tol

		for box in self.boxes:
			
			# get the list of crocs
			# cuts = list()
			# for c in self.crocs:
			#	if (c.x2 == box.ax and c.y>box.bx and c.y<box.by):
			#		cuts.append(c)
		
			# add the crocs at 15cm
			low = box.extend_crocs(self.crocs, box.ay+q, box.ay+q+t)
			if (not low):
				print("NEW", box.ax, box.bx, box.ay+q)
				low.append(Croc(box.ax,box.bx,box.ay+q))
				self.crocs = self.crocs + low
			inf = max([c.y for c in low])

			# add the crocs at -15cm
			high = box.extend_crocs(self.crocs, box.by - q, box.by - q - t)
			if (not high):
				high.append(Croc(box.ax,box.bx,box.by-q))
				self.crocs = self.crocs + high
			sup = min([c.y for c in high])
			
			print("----------------")
			# for c in self.crocs:
			#	print(c.x1, c.x2, c.y)


	def simple_crocs(self):

		q = croc_first
		p = croc_second

		for box in self.boxes:
			# crocs at +15cm and -15cm
			self.crocs.append(Croc(box.ax, box.bx, box.by - q))
			self.crocs.append(Croc(box.ax, box.bx, box.ay + q))

			# crocs at +60cm and -60cm
			if (box.by-box.ay - 2*q > 3*p):
				self.crocs.append(Croc(box.ax, box.bx, box.by - p))
				self.crocs.append(Croc(box.ax, box.bx, box.ay + p))
				
				# crocs every 120cm
				for lev in spread(box.ay+p, box.by-p, croc_maxd):
					self.crocs.append(Croc(box.ax, box.bx, lev))

			else:
				H = (box.by - box.ay - 2*q)/3
				self.crocs.append(Croc(box.ax, box.bx, box.ay + q + H))
				self.crocs.append(Croc(box.ax, box.bx, box.ay + q + 2*H))

	def total_crocs(self):
		total = 0
		for croc in self.crocs:
			total += abs(croc.x2 - croc.x1)
		return total*scale

	def get_omegas(self):	

		xc = self.xcoord
		xo = center(xc[0], xc[-1], space_omega)
		for x in xo:
			ints = intersects(self.points, x)
			ilen = len(ints)
			if (ilen > 0 and (ilen%2)==0):
				for i in range(0,ilen,2):
					self.omegas.append((x,ints[i], ints[i+1]))


	def total_omegas(self):
		total = 0
		for omega in self.omegas:
			total += abs(omega[2] - omega[1])
		return total*scale

	
	# Drawing Room

	def draw_crocs(self, msp, orient):
		for c in self.crocs:
			if (orient==0):
				line = msp.add_line((c.x1,c.y),(c.x2,c.y))
			else:
				line = msp.add_line((c.y,c.x1),(c.y,c.x2))

			line.dxf.layer = layer_croc

	def draw_omegas(self, msp, orient):
		for o in self.omegas:
			if (orient==0):
				line = msp.add_line((o[0],o[1]), (o[0],o[2]))
			else:
				line = msp.add_line((o[1],o[0]), (o[2],o[0]))
			line.dxf.layer = layer_omega
		

	def draw_room(self, msp):
		for box in self.boxes:
			box.draw_box(msp, self.orient)
			if (len(self.boxes)>1):
				box.draw_text(msp, self.index, box.number, self.orient)
			else:
				box.draw_text(msp, self.index, -1, self.orient)
			
		self.draw_crocs(msp, self.orient)
		self.draw_omegas(msp, self.orient)

	# Reporting Room

	def report_boxes(self):
		l = []
		for box in self.boxes:
			l.append(scale*scale*box.area())
		return l

	def report_crocs(self):
		l = []
		for croc in self.crocs:
			l.append(scale*croc.len())
		return l

	def report_omegas(self):
		l = []
		for omega in self.omegas:
			lomg = abs(omega[2]-omega[1])/100
			l.append(scale*lomg)
		return l

	def report(self):
		rboxes  = self.report_boxes()
		rcrocs  = self.report_crocs()
		romegas = self.report_omegas()
		return [rboxes, rcrocs, romegas]


class App:

	def __init__(self):
		self.loaded = False

		self.root = Tk()
		root = self.root
		#root.geometry('500x300')
		root.title("Eurotherm planner")
		root.resizable(width=False, height=False)

		# Control section
		ctlname = Label(root, text="Control")
		ctlname.grid(row=0, column=0, padx=(25,0), pady=(10,0), sticky="w")

		ctl = Frame(root)
		self.ctl = ctl
		ctl.config(borderwidth=1, relief='ridge')
		ctl.grid(row=1, column=0, padx=(25,25), pady=(0,20))

		button = Button(ctl, text="Open", width=5, command=self.search)
		button.grid(row=1, column=1, sticky="e")

		self.text = StringVar()
		flabel = Label(ctl, textvariable=self.text, width=30, anchor="w")
		flabel.config(borderwidth=1, relief='solid')
		flabel.grid(row=1, column=0, padx=(10,30), pady=(20,10))

		self.text1 = StringVar()
		flabel = Label(ctl, textvariable=self.text1, width=30, anchor="w")
		flabel.config(borderwidth=1, relief='solid')
		flabel.grid(row=2, column=0, padx=(10,30), pady=(10,20))

		self.var = StringVar() # variable for select layer menu

		self.button1 = Button(ctl, text="Build", width=5, command=self.build_model, pady=5)
		self.button1.grid(row=3, column=1, pady=(30,10))

		# Parameters section
		parname = Label(root, text="Settings")
		parname.grid(row=2, column=0, padx=(25,0), pady=(1,0), sticky="w")
		self.params = params = Frame(root)
		params.config(borderwidth=1, relief='ridge')
		params.grid(row=3, column=0, sticky="ew", padx=(25,25), pady=(0,2))

		Label(params, text="scale (100=1m)").grid(row=0, column=0, sticky="w")
		self.entry1 = Entry(params, justify='right', width=10)
		self.entry1.grid(row=0, column=1, sticky="w")
		self.entry1.insert(END, str(default_scale))

		Label(params, text="zone cost (m2)").grid(row=1, column=0, sticky="w")
		self.entry2 = Entry(params, justify='right', width=10)
		self.entry2.grid(row=1, column=1)
		self.entry2.insert(END, str(default_zone_cost))

		Label(params, text="transversal cuts").grid(row=2, column=0, sticky="e")
		self.tcuts = IntVar()
		self.entry3 = Checkbutton(params, variable=self.tcuts)
		self.entry3.grid(row=2, column=1)

		# Info Section
		ctlname = Label(root, text="Report")
		ctlname.grid(row=4, column=0, padx=(25,0), pady=(10,0), sticky="w")

		self.textinfo = Text(root, height=10, width=58)
		self.textinfo.config(borderwidth=1, relief='ridge')
		self.textinfo.grid(row=5, column=0, pady=(0,15)) 
		#sb = Scrollbar(root, command=self.textinfo.yview)
		#sb.grid(row=3, column=1, sticky="nsw")


		self.reset()
		self.root.mainloop()
	

	def search(self,event=None):
		self.filename = filedialog.askopenfilename(filetypes=[("DXF files", "*.dxf")])
		self.loadfile()

	def reset(self):
		self.text.set("Select DXF File")
		self.text1.set("Modified DXF File")
		self.button1["state"] = "disabled"
		self.textinfo.delete('1.0', END)

		if (hasattr(self,'opt')): self.opt.destroy()	
		self.var.set("Select layer")
		self.opt = OptionMenu(self.ctl, self.var,['Select layer'])
		self.opt.config(width=26)
		self.opt.grid(row=3, column=0, padx=(10,40), sticky="w")

	def loadfile(self):
		try:
			self.doc = ezdxf.readfile(self.filename)
		except IOError:
			self.reset()
			return
		except ezdxf.DXFStructureError:
			self.textinfo.insert(END, 'Invalid or corrupted DXF file.')
			self.reset()
			return

		self.loaded = True

		self.text.set(self.filename)
		self.outname = self.filename[:-4]+"_sapp.dxf"
		self.text1.set(self.outname)

		layers = [layer.dxf.name for layer in self.doc.layers]
		sel = layers[0]
		for layer in layers:
			if (layer == default_input_layer):
				sel = default_input_layer
				break

		self.opt.destroy()
		self.var.set(sel)
		self.opt = OptionMenu(self.ctl,self.var,*layers)
		self.opt.config(width=26)
		self.opt.grid(row=3, column=0, padx=(10,40), sticky="w")
		self.button1["state"] = "normal" 

	def new_layer(self, layer_name, color):
		attr = {'linetype': 'CONTINUOUS', 'color': color}
		self.doc.layers.new(name=layer_name, dxfattribs=attr)
		

	def create_layers(self):
		self.new_layer(layer_text, text_color)
		self.new_layer(layer_box, box_color)
		self.new_layer(layer_croc, croc_color)
		self.new_layer(layer_omega, omega_color)


	def print_report(self, txt):
		txt(END, "Design Report ----------------\n\n")
		
		for room in self.rooms:

			if (len(room.errorstr)>0):
				continue

			rep = room.report()
			boxes, crocs, omegas = rep[0], rep[1], rep[2]
			surf = room.surf * scale * scale
			txt(END,"Zone%d  - surface %.3f m2\n" % (room.index, surf))
			txt(END,"   %3d boxes, "   % len(boxes))
			txt(END," area=%8.2f m2 \n" % sum(boxes))
			txt(END,"   %3d crocs, "   % len(crocs))
			txt(END,"  tot=%8.2f m\n"   % sum(crocs))
			txt(END,"   %3d omegas, "  % len(omegas))
			txt(END," tot=%8.2f m\n"   % sum(omegas))
			txt(END,"\n")

	def save_crocs_xls(self, ws):

		sc = 66
		index = 3

		s = str(index)
		ps = [chr(sc)+s, chr(sc+1)+s, chr(sc+2)+s, chr(sc+3)+s]
		ws.column_dimensions[chr(sc)].width = 15
		ws.column_dimensions[chr(sc+1)].width = 15
		ws.column_dimensions[chr(sc+2)].width = 15
		ws.column_dimensions[chr(sc+3)].width = 15
		
		# header 
		s = str(index)
		ps = [chr(sc)+s, chr(sc+1)+s, chr(sc+2)+s, chr(sc+3)+s]
		ws[chr(sc)+s].value = 'Crocodiles'
		index += 1

		s = str(index)
		ps = [chr(sc)+s, chr(sc+1)+s, chr(sc+2)+s, chr(sc+3)+s]
		ws[ps[0]].value = "Zone name"
		ws[ps[0]].border = Border(top=Side(style='thin'), bottom=Side(style='double'),
								 left=Side(style='thin')) 
		ws[ps[1]].value = "total length (m)"
		ws[ps[1]].border = Border(top=Side(style='thin'), bottom=Side(style='double')) 
		ws[ps[2]].value = "No. profiles"
		ws[ps[2]].border = Border(top=Side(style='thin'), bottom=Side(style='double')) 
		ws[ps[3]].value = "No. packages"
		ws[ps[3]].border = Border(top=Side(style='thin'), bottom=Side(style='double'),
								 right=Side(style='thin')) 
		index += 1

		total_len = total_profs = total_packs = 0 
		for room in self.rooms:

			s = str(index)
			ps = [chr(sc)+s, chr(sc+1)+s, chr(sc+2)+s, chr(sc+3)+s]
			ws[ps[0]].border = Border(left=Side(style='thin')) 
			ws[ps[3]].border = Border(right=Side(style='thin')) 
			ws[ps[0]].value = "Zone " + str(room.index)
			index += 1
			if (room.ignore):
				continue

			ws[ps[1]].value = tot = room.total_crocs()/100
			total_len += tot
			ws[ps[1]].number_format = "0.00"

		s = str(index)
		ps = [chr(sc)+s, chr(sc+1)+s, chr(sc+2)+s, chr(sc+3)+s]
		ws[ps[0]].border = Border(bottom=Side(style='thin'), top=Side(style='double'), 
									left=Side(style='thin')) 
		ws[ps[1]].border = Border(bottom=Side(style='thin'), top=Side(style='double')) 
		ws[ps[2]].border = Border(bottom=Side(style='thin'), top=Side(style='double')) 
		ws[ps[3]].border = Border(bottom=Side(style='thin'), top=Side(style='double'),
								 right=Side(style='thin')) 

		ws[ps[0]].value = "totals"
		ws[ps[1]].value = total_len
		ws[ps[1]].number_format = "0.00"
		ws[ps[2]].value = tot = ceil(total_len/5)
		ws[ps[3]].value = ceil(tot/10)

	def save_omegas_xls(self, ws):
		sc = 72
		index = 3

		s = str(index)
		ps = [chr(sc)+s, chr(sc+1)+s, chr(sc+2)+s, chr(sc+3)+s]
		ws.column_dimensions[chr(sc)].width = 15
		ws.column_dimensions[chr(sc+1)].width = 15
		ws.column_dimensions[chr(sc+2)].width = 15
		ws.column_dimensions[chr(sc+3)].width = 15
		
		# header 
		s = str(index)
		ps = [chr(sc)+s, chr(sc+1)+s, chr(sc+2)+s, chr(sc+3)+s]
		ws[chr(sc)+s].value = 'Omegas'
		index += 1

		s = str(index)
		ps = [chr(sc)+s, chr(sc+1)+s, chr(sc+2)+s, chr(sc+3)+s]
		ws[ps[0]].value = "Zone name"
		ws[ps[0]].border = Border(top=Side(style='thin'), bottom=Side(style='double'),
								 left=Side(style='thin')) 
		ws[ps[1]].value = "total length (m)"
		ws[ps[1]].border = Border(top=Side(style='thin'), bottom=Side(style='double')) 
		ws[ps[2]].value = "No. profiles"
		ws[ps[2]].border = Border(top=Side(style='thin'), bottom=Side(style='double')) 
		ws[ps[3]].value = "No. packages"
		ws[ps[3]].border = Border(top=Side(style='thin'), bottom=Side(style='double'),
								 right=Side(style='thin')) 
		index += 1

		total_len = 0 
		for room in self.rooms:

			s = str(index)
			ps = [chr(sc)+s, chr(sc+1)+s, chr(sc+2)+s, chr(sc+3)+s]
			index += 1
			ws[ps[0]].border = Border(left=Side(style='thin')) 
			ws[ps[3]].border = Border(right=Side(style='thin')) 
			ws[ps[0]].value = "Zone " + str(room.index)

			if (room.ignore):
				continue

			ws[ps[1]].value = tot = room.total_omegas()/100
			ws[ps[1]].number_format = "0.00"
			total_len += tot

		s = str(index)
		ps = [chr(sc)+s, chr(sc+1)+s, chr(sc+2)+s, chr(sc+3)+s]
		ws[ps[0]].border = Border(bottom=Side(style='thin'), top=Side(style='double'), 
									left=Side(style='thin')) 
		ws[ps[1]].border = Border(bottom=Side(style='thin'), top=Side(style='double')) 
		ws[ps[2]].border = Border(bottom=Side(style='thin'), top=Side(style='double')) 
		ws[ps[3]].border = Border(bottom=Side(style='thin'), top=Side(style='double'),
								 right=Side(style='thin')) 

		ws[ps[0]].value = "totals"
		ws[ps[1]].value = total_len
		ws[ps[1]].number_format = "0.00"
		ws[ps[2]].value = tot = ceil(total_len/5)
		ws[ps[3]].value = ceil(tot/10)
		

	def save_xls(self):
		wb = openpyxl.load_workbook(xlsx_template)
		ws = wb[sheet_template]

		index = 0
		for room in self.rooms:

			if (len(room.errorstr)>0):
				curr_row = str(index+6)
				index = index + 1
				pos_name = 'B' + curr_row
				pos_err = 'C' + curr_row
				ws[pos_name] = 'Zone ' + str(room.index)
				continue

			for box in room.boxes:
				curr_row = str(index+6)
				index = index + 1

				pos_name = 'B' + curr_row
				pos_paneltype = 'C' + curr_row
				pos_start_profile = 'D' + curr_row
				pos_len = 'E' + curr_row
				pos_end_profile = 'F' + curr_row
				pos_width = 'G' + curr_row
				pos_perimeter = 'H' + curr_row
				pos_omega = 'I' + curr_row

				if (len(room.boxes) == 1):
					ws[pos_name] = "Zone %d" % (room.index) 
				else:
					ws[pos_name] = "Zone %d " % (room.index) + chr(64+box.number)
					perimeter = 0.0

				if (box.number == 1):
					perimeter = room.perimeter()*scale/100
				else:
					perimenter = 0.0

				width = (box.bx - box.ax)*scale
				length = (box.by - box.ay)*scale + extra_len

				ws[pos_width] = ceil(width/5)*5/100
				ws[pos_paneltype] = 'BLH'
				ws[pos_start_profile] = 'NONE'
				ws[pos_len] = round(length/5)*50
				ws[pos_start_profile] = 'NONE'
				ws[pos_end_profile] = 'NONE'
				ws[pos_perimeter] = perimeter
				ws[pos_omega] = '0.0'

				#ws[pos_area].number_format = "0.00"

		out = self.filename[:-4] + "_doghe.xlsx"	
		wb.save(out)


	def build_model(self):
		global width_doga, space_omega  
		global croc_first, croc_second, croc_maxd, croc_tol     
		global cut_size, zone_cost   
		global x_font_size, y_font_size  
		global scale, tolerance
		global tcuts

		self.textinfo.delete('1.0', END)

		if (not self.loaded):
			self.textinfo(END, "File not loaded")
			return

		scale = float(self.entry1.get())
		zone_cost = float(self.entry2.get())

		tolerance    = default_tolerance/scale
		width_doga   = default_width_doga/scale
		space_omega  = default_space_omega/scale
		croc_first   = default_croc_first/scale
		croc_second  = default_croc_second/scale
		croc_maxd    = default_croc_maxd/scale
		croc_tol     = default_croc_tol/scale
		cut_size     = default_cut_size/scale
		zone_cost    = zone_cost/(scale*scale)
		x_font_size  = default_x_font_size/scale
		y_font_size  = default_y_font_size/scale

		# reload file
		self.doc = ezdxf.readfile(self.filename)	
		self.msp = self.doc.modelspace()
		Room.index = 1
		self.rooms = []

		self.create_layers()
		inputlayer = self.var.get()

		for e in self.msp.query('*[layer=="%s"]' % inputlayer):
			if (e.dxftype() != 'LWPOLYLINE'):
				wstr = "WARNING: layer contains non-polyline: %s\n" % e.dxftype()
				self.textinfo.insert(END, wstr)

		searchstr = 'LWPOLYLINE[layer=="'+inputlayer+'"]'
		query = self.msp.query(searchstr)
		if (len(query) == 0):
			wstr = "WARNING: layer %s does not contain polylines\n" % inputlayer
			self.textinfo.insert(END, wstr)

		tcuts = self.tcuts.get()

		for poly in query:
			room = Room(poly)
			room.surf = 0
			self.rooms.append(room)
			if (len(room.errorstr)>0):
				self.textinfo.insert(END,room.errorstr)
			else:
				room.surf = room.area()
				if (room.area() < zone_cost):
					wstr = "WARNING: area less than %d m2: " % zone_cost
					wstr += "Consider changing scale!\n"
					self.textinfo.insert(END, wstr)
				room.draw_room(self.msp)

		self.print_report(self.textinfo.insert)

		if (os.path.isfile(self.outname)):
			if askyesno("Warning", "File 'mod' already exists: Overwrite?"):
				self.doc.saveas(self.outname)
		else:
				self.doc.saveas(self.outname)

		self.save_xls()
		wb = openpyxl.Workbook()
		ws = wb.active
		ws.title = "Bill of Materials"
		if (len(self.rooms)>0):
			self.save_crocs_xls(ws)
			self.save_omegas_xls(ws)
		out = self.filename[:-4] + "_struct.xlsx"	
		wb.save(out)

	
App()


