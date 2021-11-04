#!/usr/local/bin/python3

import ezdxf
import openpyxl
from math import ceil, floor, sqrt
from tkinter import *
from tkinter import filedialog

# Parameter settings (values in cm)
tolerance    = 1         # ignore too little variations
width_doga   = 20
space_omega  = 60
croc_first   = 15
croc_second  = 60
croc_maxd    = 120
croc_tol     = 10

zone_cost    = 0.5     # equivalent cost of a zone in m2 material

x_font_size  = 20
y_font_size  = 30

default_input_layer = 'aree sapp'
layer_text   = 'Eurotherm_text'
layer_box    = 'Eurotherm_box'
layer_croc   = 'Eurotherm_crocodile'
layer_omega   = 'Eurotherm_omega'


alphabet = {
	' ': [],
	'A': [12,3,1,5,14,8,6],
	'B': [0,1,5,7,6,7,11,13,12,0],
	'C': [2,1,3,9,13,14],
	'D': [0,1,5,11,13,12,0],
	'E': [2,0,6,7,6,12,14],
	'F': [2,0,6,7,6,12],
	'G': [2,1,3,9,13,14,8,7],
	'H': [0,12,6,8,2,14],
	'I': [1,13],
	'J': [9,13,1,0,1,2],
	'K': [0,6,2,6,12,6,14],
	'L': [0,12,14],
	'M': [12,0,7,2,14],
	'N': [12,0,14,2],
	'O': [3,1,5,11,13,9,3],
	'P': [12,0,1,5,7,6],
	'Q': [13,11,5,1,3,9,13,14,13,10],
	'R': [12,0,1,5,7,6,7,11,14],
	'S': [12,13,11,3,1,2],
	'T': [0,2,1,13],
	'U': [0,9,13,11,2],
	'V': [0,13,2],
	'W': [0,13,2],
	'X': [0,14,7,2,12],
	'Y': [0,7,2,7,13],
	'Z': [0,2,12,14],
	'0': [1,3,9,13,11,5,1],
	'1': [3,1,13,12,14],
	'2': [3,1,5,12,14],
	'3': [0,1,5,7,6,7,11,13,12],
	'4': [0,6,8,2,14],
	'5': [2,0,6,7,11,13,12],
	'6': [2,1,3,9,13,11,7,6],
	'7': [0,2,12],
	'8': [3,1,5,9,12,14,11,3],
	'9': [8,6,3,1,5,11,13,12]
}

def pletter(letter,pos,scale):
	path = alphabet[letter]
	poly = []
	for p in path:
		x = p % 3
		y = 4 - (p//3)
		poly.append((pos[0]+x*scale[0], pos[1]+y*scale[1]))
	return poly

def writedxf(msp, strn, pos, scale):

	slen = len(strn)*scale[0]*3
	pos = (pos[0]-slen/2, pos[1]-scale[1]*2)
	
	for l in strn:
		poly = pletter(l, pos, scale)
		pl = msp.add_lwpolyline(poly)
		pl.dxf.layer = layer_text
		pos = (pos[0]+scale[0]*3, pos[1])

def spread(a,b,size):
	l = []
	n = ceil(abs(a-b)/size) - 1
	L = abs(a-b)/(n+1)
	for i in range(0,n):
		l.append((i+1)*L+a)
	return l

def center(a,b,size):
	l = []
	n = ceil(abs(b-a)/size) - 1
	L = (abs(b-a) - size*n)/2
	for i in range(0,n+1):
		l.append(a+L+i*size)	
	return l


class Split:
	
	def __init__(self, ax, bx, lu, ld, ru, rd):
		self.ax = ax
		self.bx = bx
		self.lu = lu
		self.ld = ld
		self.ru = ru
		self.rd = rd

	def base(self):
		return abs(self.bx - self.ax)

	def area(self):
		l1 = abs(self.lu[0] - self.ld[0])
		l2 = abs(self.ru[0] - self.rd[0])
		b = self.base()
		return (l1+l2)*b/20000


class Cluster:
	def __init__(self, box, x):
		self.x = x
		self.ax = box.ax
		self.bx = box.bx
		self.ay = box.ay
		self.by = box.by
		if (box.bx == x):
			self.boxesl = [box]
			self.boxesr = []
		else:
			self.boxesl = []
			self.boxesr = [box]

	def print(self):

		print("x=%8.4f [%d %d %d %d]:" % (self.x, self.ax, self.bx, self.ay, self.by))

		print("LB:", end='')
		for box in self.boxesl:
			print("box(%d) [%d %d %d %d]   " % (box.id, box.ax, box.bx, box.ay, box.by), end='')
		print("")

		print("RB:", end='')
		for box in self.boxesr:
			print("box(%d) [%d %d %d %d]   " % (box.id, box.ax, box.bx, box.ay, box.by), end='')
		print("\n")

		
	def bb_area(self):
		ax = self.ax
		ay = self.ay
		bx = self.bx
		by = self.by
		return abs(bx-ax)*abs(by-ay)/10000

	def area(self):
		area_ls = 0
		for box in self.boxesl:
			area_ls += box.area()
		area_rs = 0
		for box in self.boxesr:
			area_rs += box.area()

		return area_ls + area_rs
		
	def scrap(self):
		s =  self.bb_area() - self.area()
		return s
		

	def append(self, box):

		if (box.ax != self.x and box.bx != self.x):
			return

		if (box.ax < self.ax):
			self.ax = box.ax

		if (box.bx > self.bx):
			self.bx = box.bx

		if (box.ay < self.ay):
			self.ay = box.ay

		if (box.by > self.by):
			self.by = box.by

		if (box.ax == self.x):
			self.boxesr.append(box)
		
		if (box.bx == self.x):
			self.boxesl.append(box)


	def check_append(self, box):
		if (box.ax==self.x or box.bx==self.x):
			if ((box.ay>self.ay and box.ay<self.by) or
				(box.by>self.ay and box.by<self.by)):
				self.append(box)
				return True
		return False

	def can_merge(self, cluster):
		if (self.x == cluster.x):
			if (self.ay<self.by or self.by>self.ay):
				return True
		return False

	@classmethod
	def _get_one(cls, clusters):
		flag = False
		for cl1 in clusters:
			for cl2 in clusters:
				if (cl2==cl1):
					continue
				if (cl2.can_merge(cl1)):
					return (cl1, cl2)

	@classmethod
	def merge(cls, clusters):
		
		while( (cc := cls._get_one(clusters)) != None):

			cc[0].boxesr += cc[1].boxesr
			cc[0].boxesl += cc[1].boxesl

			if (cc[0].ay > cc[1].ay):
				cc[0].ay = cc[1].ay
			if (cc[0].by < cc[1].by):
				cc[0].by = cc[1].by
			if (cc[0].ax > cc[1].ax):
				cc[0].ax = cc[1].ax
			if (cc[0].bx < cc[1].bx):
				cc[0].bx = cc[1].bx

			clusters.remove(cc[1])


class Croc:
	def __init__(self, x1, x2, y):
		self.x1 = x1
		self.x2 = x2
		self.y  = y 

	def fit(self, y1, y2):
		tol = croc_tol
		return (self.y>=y1 and self.y<=y2)


	def extend_right(self, x2):
		self.x2 = x2

	def len(self):
		return abs(self.x2-self.x1)/100

class Zone:
	
	ident = 0

	def __init__(self, ax, ay, bx, by, split):
		self.id = Zone.ident
		Zone.ident += 1
		self.ax = min(ax, bx)
		self.bx = max(ax, bx)
		self.ay = min(ay, by)
		self.by = max(ay, by)
		self.TR = (max(ax, bx), max(ay, by))
		self.TL = (min(ax, bx), max(ay, by))
		self.BR = (max(ax, bx), min(ay, by))
		self.BL = (min(ax, bx), min(ay, by))
		if (type(split) == list):
			self.splits = split
		else:
			self.splits = [split]

	def pline(self, orient):
		ax = self.ax; bx = self.bx
		ay = self.ay; by = self.by
		if (orient == 0):
			return [(ax,ay),(ax,by),(bx,by),(bx,ay),(ax,ay)]
		else:	
			return [(ay,ax),(by,ax),(by,bx),(ay,bx),(ay,ax)]

	def extend_right(self, coord):
		self.TR = (coord, self.TR[1])
		self.BR = (coord, self.BR[1])
		self.bx = coord

	def extend_crocs(self, crocs, y1, y2):
		crocs_list = list()
		for c in crocs:
			print("BOX",self.ax,y1,y2)
			if (c.x2 == self.ax and c.y>=y1 and c.y<=y2):
				print("exte",c.x1,c.x2,c.y)
				c.extend_right(self.bx)
				crocs_list.append(c)
		return crocs_list

	def draw_box(self, msp, orient):
		pl = msp.add_lwpolyline(self.pline(orient))
		pl.dxf.layer = layer_box

	def draw_text(self, msp, ind, subind, orient):
		scale = (x_font_size/2, y_font_size/4)
		text = 'ZONE' + str(ind) + chr(65+subind%26) 
		if (orient==0):
			pos = ((self.ax+self.bx)/2, (self.ay+self.by)/2)
		else:
			pos = ((self.ay+self.by)/2, (self.ax+self.bx)/2)
	
		writedxf(msp, text, pos, scale)

	def area(self):
		return (self.bx-self.ax)*(self.by-self.ay)/10000

	def splitarea(self):
		totsplit = 0
		for split in self.splits:
			totsplit = totsplit + split.area()
		return totsplit

	def scrap(self):
		return self.area() - self.splitarea()



class Room:

	index = 0

	def __init__(self, poly):


		self.errorstr = ""

		tol = tolerance
		self.orient = 0
		self.boxes = list()
		self.coord = list()
		self.crocs = list()
		self.omegas = list()

		self.index = Room.index
		Room.index = Room.index + 1

		self.points = list(poly.vertices())

		# Add a final point to closed polylines
		p = self.points
		if (poly.is_closed):
			self.points.append((p[0][0], p[0][1]))

		# Check if the polyine is open with large final gap
		n = len(p)-1
		if (abs(p[0][0]-p[n][0])>tol or abs(p[0][1]-p[n][1])>tol):
			self.errorstr = "Error: open polyline in Zone %d \n" % self.index 
			return

		# Straighten up walls
		finished = False
		while not finished:
			for i in range(1, len(p)):
				if (abs(p[i][0]-p[i-1][0]) < tol):
					p[i] = (p[i-1][0], p[i][1])

				if (abs(p[i][1]-p[i-1][1]) < tol):
					p[i] = (p[i][0], p[i-1][1])

			if (p[0] == p[n]):
				finished = True
			else:
				p[0] = p[n]

		#for i in range(1, len(p)):
		#	print(abs(p[i][0]-p[i-1][0]), abs(p[i][1]-p[i-1][1]))

		# Projections of coordinates on x and y
		self.xcoord = sorted(set([p[0] for p in self.points]))	
		self.ycoord = sorted(set([p[1] for p in self.points]))	

		# Mirror if polyline is green
		if (poly.dxf.color==3):
			self.orient = 1
			for i in range(0,len(self.points)):
				self.points[i] = (self.points[i][1], self.points[i][0])
			self.xcoord, self.ycoord = self.ycoord, self.xcoord

		self.get_boxes()
		#self.get_crocs()
		self.simple_crocs()
		self.get_omegas()


	# Building Room

	# This function divides a polyline into boxes
	# and returns the list of boxes
	def get_boxes(self):
		
		tol = tolerance
		boxes = self.boxes
		p = self.points
		xc = self.xcoord

		for i in range(0,len(xc)-1):
			ax = xc[i]
			bx = xc[i+1]
			mid = (ax + bx)/2

			lr = list()
			ll = list()

			for j in range(0, len(p)-1):
				if (p[j][0] < p[j+1][0]):
					x0, y0 = p[j][0], p[j][1]
					x1, y1 = p[j+1][0], p[j+1][1]
				else:
					x0, y0 = p[j+1][0], p[j+1][1]
					x1, y1 = p[j][0], p[j][1]
					
				if (x0<mid and mid<x1):
					delta = x1 - x0
					ay = (y0*(x1-ax) - y1*(x0-ax))/delta
					by = (y0*(x1-bx) - y1*(x0-bx))/delta
					ll.append((ay,j))
					lr.append((by,j))

			ll.sort()
			lr.sort()

			for j in range(0,len(ll),2):
				ay = min(ll[j][0], lr[j][0])
				by = max(ll[j+1][0], lr[j+1][0])

				split = Split(ax, bx, ll[j], ll[j+1], lr[j], lr[j+1])
				boxes.append(Zone(ax,ay,bx,by, split))

		# collate splits
		while (self.collate_splits()): pass

		# minimize scrap
		to_add = []
		to_del = []
		for box in boxes:
			area = box.area()
			scrap = box.scrap()
			m = sqrt(scrap/zone_cost)
			n = max(floor(m), 1)		
			if ( n*zone_cost < scrap/(n+1) ):
				n = n + 1

			if (n>1):
				# instantiate new blocks
				ax = box.ax
				bx = box.bx
				d = (bx - ax)/n
				lu = box.splits[0].lu[0]
				ld = box.splits[0].ld[0]
				du = (box.splits[0].ru[0] - lu)/n
				dd = (box.splits[0].rd[0] - ld)/n
				
				j = box.splits[0].lu[1]
				k = box.splits[0].ld[1]

				for i in range(0,n):
					ru = lu + du
					rd = ld + dd
					split = Split(ax, ax+d, (lu,j), (ld,k), (ru,j), (rd,k))
					ay = min(lu, ld, ru, rd)
					by = max(lu, ld, ru, rd)
					zs = Zone(ax, ay, ax + d, by, split)
					to_add.append(zs)
					ax = ax + d
					lu = lu + du
					ld = ld + dd 

				to_del.append(box)

		self.boxes = boxes + to_add
		for box in to_del:
			self.boxes.remove(box)


		# optimization 

		# collapse clusters
		# rounds = 0
		done = False
		while(not done):
			self.get_clusters()
			cost = 2*zone_cost
			for c in self.clusters:
				nl = len(c.boxesl)
				nr = len(c.boxesr)
				if (nl>0 and nr>0):
					#print("func:", "R: %d L: %d" % (nl, nr), 
					#		   "%.4f" % c.bb_area(), 
					#		   "%.4f" % c.area(), 
					#		   "%.4f" % c.scrap())
					scrap = c.scrap()
					if (scrap < cost):
						cost = scrap
						best = c

			if (cost<zone_cost):

				#print(  " >>>>>> found best", end='')
				#print("(%d %d %d %d)" % (best.ax, best.bx, best.ay, best.by))
				#print("left boxes: ", end='')
				#for box in best.boxesl:
				#	print("box(%d) " % box.id, end='')
				#	print("[%d %d %d %d]" % (box.ax, box.bx, box.ay, box.by), end='')
				#print("")
				#print("right boxes: ", end='')
				#for box in best.boxesr:
				#	print("box(%d) " % box.id, end='')
				#	print("[%d %d %d %d]" % (box.ax, box.bx, box.ay, box.by), end='')
				#print("\n<<<<<")

				splits = []
				for box in best.boxesl:
					splits += box.splits
				for box in best.boxesr:
					splits += box.splits

				zone = Zone(best.ax, best.ay, best.bx, best.by, splits)
				self.boxes.append(zone)

				for box in best.boxesl:
					self.boxes.remove(box)

				for box in best.boxesr:
					self.boxes.remove(box)

			else:
				done = True

	def print_boxes(self):
		print("LIST OF %d BOXES: " % len(self.boxes), end='')
		for box in self.boxes:
			print("[%d]" % box.id, end='')
		print("")
			

	def print_clusters(self):		
		print("START CLUSTERS -------------------------")
		for cluster in self.clusters:
			cluster.print()
		print('END CLUSTERS   =========================')

	def get_clusters(self):
		clusters = self.clusters = []
		for box in self.boxes:

			fl = False
			fr = False
			for cluster in clusters:
				if (cluster.check_append(box)):
					if (box.ax == cluster.x):
						fl = True
					if (box.bx == cluster.x):
						fr = True

			if (fl == False):
				clusters.append(Cluster(box, box.ax))

			if (fr == False):
				clusters.append(Cluster(box, box.bx))

		Cluster.merge(clusters)
				

	def collate_splits(self):
		for b in self.boxes:
			for q in self.boxes:
				if (b.ax == q.bx and b.splits[0].lu[1] == q.splits[0].ru[1] and
					                 b.splits[0].ld[1] == q.splits[0].rd[1]):
					lu = q.splits[0].lu 
					ld = q.splits[0].ld
					ru = b.splits[0].ru
					rd = b.splits[0].rd
					split = Split(q.ax, b.bx, lu, ld, ru, rd)
					top = max(q.by, b.by)
					btm = min(q.ay, b.ay)
					z = Zone(q.ax, btm, b.bx, top, split)
					self.boxes.remove(b)
					self.boxes.remove(q)
					self.boxes.append(z)
					return True

		return False


	def get_crocs(self):
		
		q = croc_first
		t = croc_tol

		for box in self.boxes:
			
			# get the list of crocs
			# cuts = list()
			# for c in self.crocs:
			#	if (c.x2 == box.ax and c.y>box.bx and c.y<box.by):
			#		cuts.append(c)
		
			# add the crocs at 15cm
			low = box.extend_crocs(self.crocs, box.ay+q, box.ay+q+t)
			if (not low):
				print("NEW", box.ax, box.bx, box.ay+q)
				low.append(Croc(box.ax,box.bx,box.ay+q))
				self.crocs = self.crocs + low
			inf = max([c.y for c in low])

			# add the crocs at -15cm
			high = box.extend_crocs(self.crocs, box.by - q, box.by - q - t)
			if (not high):
				high.append(Croc(box.ax,box.bx,box.by-q))
				self.crocs = self.crocs + high
			sup = min([c.y for c in high])
			
			print("----------------")
			# for c in self.crocs:
			#	print(c.x1, c.x2, c.y)


	def simple_crocs(self):

		q = croc_first
		p = croc_second

		for box in self.boxes:
			# crocs at +15cm and -15cm
			self.crocs.append(Croc(box.ax, box.bx, box.by - q))
			self.crocs.append(Croc(box.ax, box.bx, box.ay + q))

			# crocs at +60cm and -60cm
			if (box.by-box.ay - 2*q > 3*p):
				self.crocs.append(Croc(box.ax, box.bx, box.by - p))
				self.crocs.append(Croc(box.ax, box.bx, box.ay + p))
				
				# crocs every 120cm
				for lev in spread(box.ay+p, box.by-p, croc_maxd):
					self.crocs.append(Croc(box.ax, box.bx, lev))

			else:
				H = (box.by - box.ay - 2*q)/3
				self.crocs.append(Croc(box.ax, box.bx, box.ay + p + H))
				self.crocs.append(Croc(box.ax, box.bx, box.ay + p + 2*H))

	def get_omegas(self):	
		for box in self.boxes:			
			for d in center(box.ax, box.bx, space_omega):
				self.omegas.append((d,box.ay,box.by))

	
	# Drawing Room

	def draw_crocs(self, msp, orient):
		for c in self.crocs:
			if (orient==0):
				line = msp.add_line((c.x1,c.y),(c.x2,c.y))
			else:
				line = msp.add_line((c.y,c.x1),(c.y,c.x2))

			line.dxf.layer = layer_croc

	def draw_omegas(self, msp, orient):
		for o in self.omegas:
			if (orient==0):
				line = msp.add_line((o[0],o[1]), (o[0],o[2]))
			else:
				line = msp.add_line((o[1],o[0]), (o[2],o[0]))
			line.dxf.layer = layer_omega
		

	def draw_room(self, msp):

		subindex = 0
		for box in self.boxes:
			box.draw_box(msp, self.orient)
			# box.draw_text(msp, self.index, subindex, self.orient)
			subindex = subindex + 1
			
		# self.draw_crocs(msp, self.orient)
		# self.draw_omegas(msp, self.orient)

	# Reporting Room

	def report_boxes(self):
		l = []
		for box in self.boxes:
			l.append(box.area())
		return l

	def report_crocs(self):
		l = []
		for croc in self.crocs:
			l.append(croc.len())
		return l

	def report_omegas(self):
		l = []
		for omega in self.omegas:
			lomg = abs(omega[2]-omega[1])/100
			l.append(lomg)
		return l

	def report(self):
		rboxes  = self.report_boxes()
		rcrocs  = self.report_crocs()
		romegas = self.report_omegas()
		return [rboxes, rcrocs, romegas]



class App:

	rooms = list()

	def __init__(self):
		self.root = Tk()
		root = self.root
		#root.geometry('500x300')
		root.title("Eurotherm planner")
		root.resizable(width=False, height=False)

		# Control section
		ctlname = Label(root, text="Control")
		ctlname.grid(row=0, column=0, padx=(25,0), pady=(10,0), sticky="w")

		ctl = Frame(root)
		self.ctl = ctl
		ctl.config(borderwidth=1, relief='ridge')
		ctl.grid(row=1, column=0, padx=(25,25), pady=(0,20))

		button = Button(ctl, text="Open", width=5, command=self.search)
		button.grid(row=1, column=1, sticky="e")

		self.text = StringVar()
		self.text.set("Select DXF File")
		flabel = Label(ctl, textvariable=self.text, width=30, anchor="w")
		flabel.config(borderwidth=1, relief='solid')
		flabel.grid(row=1, column=0, padx=(10,30), pady=(20,10))

		self.text1 = StringVar()
		self.text1.set("Modified DXF File")
		flabel = Label(ctl, textvariable=self.text1, width=30, anchor="w")
		flabel.config(borderwidth=1, relief='solid')
		flabel.grid(row=2, column=0, padx=(10,30), pady=(10,20))

		self.var = StringVar()
		self.var.set("Select layer")
		self.opt = OptionMenu(ctl, self.var,['Select layer'])
		self.opt.config(width=26)
		self.opt.grid(row=3, column=0, padx=(10,40), sticky="w")

		self.button1 = Button(ctl, text="Build", width=5, command=self.build_model, pady=5)
		self.button1.grid(row=4, column=1, pady=(30,10))
		self.button1["state"] = "disabled"


		# Info Section
		ctlname = Label(root, text="Report")
		ctlname.grid(row=2, column=0, padx=(25,0), pady=(10,0), sticky="w")

		self.textinfo = Text(root, height=10, width=58)
		self.textinfo.config(borderwidth=1, relief='ridge')
		self.textinfo.grid(row=3, column=0, pady=(0,15)) 
		#sb = Scrollbar(root, command=self.textinfo.yview)
		#sb.grid(row=3, column=1, sticky="nsw")

		self.root.mainloop()
	

	def search(self,event=None):
		self.filename = filedialog.askopenfilename(filetypes=[("DXF files", "*.dxf")])
		self.loadfile()

	def loadfile(self):
		try:
			self.doc = ezdxf.readfile(self.filename)
		except IOError:
			self.text.set('Not a DXF file or a generic I/O error.')
			return
		except ezdxf.DXFStructureError:
			self.text.set('Invalid or corrupted DXF file.')
			return

		self.text.set(self.filename)
		self.outname = self.filename[:-4]+"_mod.dxf"
		self.text1.set(self.outname)

		self.msp = self.doc.modelspace()
		
		layers = [layer.dxf.name for layer in self.doc.layers]
		sel = layers[0]
		for layer in layers:
			if (layer == default_input_layer):
				sel = default_input_layer
				break

		self.opt.destroy()
		self.var.set(sel)
		self.opt = OptionMenu(self.ctl,self.var,*layers)
		self.opt.config(width=26)
		self.opt.grid(row=3, column=0, padx=(10,40), sticky="w")
		self.button1["state"] = "normal" 

	def create_layers(self):
		self.doc.layers.new(name=layer_text, dxfattribs={'linetype': 'CONTINUOUS', 'color': 7})
		self.doc.layers.new(name=layer_box, dxfattribs={'linetype': 'CONTINUOUS', 'color': 8})
		self.doc.layers.new(name=layer_croc, dxfattribs={'linetype': 'CONTINUOUS', 'color': 9})
		self.doc.layers.new(name=layer_omega, dxfattribs={'linetype': 'CONTINUOUS', 'color': 10})

	def print_report(self, txt):
		txt(END, "Design Report ----------------\n\n")
		
		for room in self.rooms:

			if (len(room.errorstr)>0):
				continue

			rep = room.report()
			boxes, crocs, omegas = rep[0], rep[1], rep[2]

			txt(END,"Zone"+str(room.index)+":\n")
			txt(END,"   %3d boxes, "   % len(boxes))
			txt(END," area=%8.2f m2 \n" % sum(boxes))
			txt(END,"   %3d crocs, "   % len(crocs))
			txt(END,"  tot=%8.2f m\n"   % sum(crocs))
			txt(END,"   %3d omegas, "  % len(omegas))
			txt(END," tot=%8.2f m\n"   % sum(omegas))
			txt(END,"\n")

	def save_xls(self):
		wb = openpyxl.Workbook()
		ws = wb.active
		ws.title = "Bills of Materials"
	
		index = 0
		for room in self.rooms:

			curr_row = str(index+4)
			curr_zone = str(index)
			index = index + 1

			pos_name = 'B' + curr_row
			pos_box  = 'C' + curr_row
			pos_croc = 'D' + curr_row
			pos_omeg = 'E' + curr_row

			ws[pos_name] = 'Zone' + curr_zone

			if (len(room.errorstr)>0):
				ws[pos_box] = "ERROR"
				continue

			rep = room.report()
			boxes, crocs, omegas = rep[0], rep[1], rep[2]

			A_tot = sum(boxes)
			crocs_tot = sum(crocs)
			omegs_tot = sum(omegas)

			ws[pos_box] = A_tot
			ws[pos_croc] = crocs_tot
			ws[pos_omeg] = omegs_tot

		out = self.filename[:-4] + "_mod.xlsx"	
		wb.save(out)


	def build_model(self):

		self.create_layers()
		inputlayer = self.var.get()
		searchstr = 'LWPOLYLINE[layer=="'+inputlayer+'"]'
		for poly in  self.msp.query(searchstr):
			room = Room(poly)
			self.rooms.append(room)
			if (len(room.errorstr)>0):
				self.textinfo.insert(END,room.errorstr)
			else:
				room.draw_room(self.msp)

		self.print_report(self.textinfo.insert)
		self.doc.saveas(self.outname)
		self.save_xls()


	
App()



